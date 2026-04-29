#!/usr/bin/env python3
"""
Jamførar eit Excel‑ark (eller CSV lagra frå Excel) med NU‑oversikt mot lagsregisteret.

Forventar minst ei kolonne med **organisasjonsnummer** (9 siffer). Valfri **status**-
kolonne (nedlagt/utmeld/medlem o.l.) jamførast med data/manuell_status.json.

Døme:
  pip install -r requirements.txt   # gir openpyxl om du brukar .xlsx
  python3 skript/jamfoer_excel_mot_register.py /sti/til/nu-oversikt.xlsx
  python3 skript/jamfoer_excel_mot_register.py data/nu_oversikt.csv \\
      --orgnr-kolonne Org.nr --status-kolonne Status
  python3 skript/jamfoer_excel_mot_register.py fil.csv --med-tom-manuell \\
      --max-utanfor-excel 0

Statusjamføring berre for orgnr som **òg finst i lag.csv**. «I CSV men ikkje i Excel»
kan vere mange rader dersom Excel‑lista berre gjeld NU — bruk --max-utanfor-excel 0
for berre tal.

Berre rapport til stdout — skriv ikkje automatisk til manuell_status (du vurderer sjølv).
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

def rot() -> Path:
    return Path(__file__).resolve().parent.parent


def krev_openpyxl():
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


def les_xlsx(sti: Path, ark: str | int | None) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(sti, read_only=True, data_only=True)
    if isinstance(ark, str) and ark in wb.sheetnames:
        ws = wb[ark]
    elif isinstance(ark, int) and 0 <= ark < len(wb.sheetnames):
        ws = wb[wb.sheetnames[ark]]
    else:
        ws = wb.active
    rader_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rader_raw:
        return [], []
    h = [str(c).strip() if c is not None else "" for c in rader_raw[0]]
    # unike overskrifter
    brukt: dict[str, int] = {}
    h2: list[str] = []
    for x in h:
        n = x or f"kol_{len(h2)}"
        if n in brukt:
            brukt[n] += 1
            n = f"{n}_{brukt[n]}"
        else:
            brukt[n] = 0
        h2.append(n)
    ut: list[dict[str, str]] = []
    for rad in rader_raw[1:]:
        d: dict[str, str] = {}
        for i, k in enumerate(h2):
            v = rad[i] if i < len(rad) else None
            if v is None:
                d[k] = ""
            elif isinstance(v, float) and v == int(v):
                d[k] = str(int(v))
            else:
                d[k] = str(v).strip()
        ut.append(d)
    return h2, ut


def les_csv(sti: Path) -> tuple[list[str], list[dict[str, str]]]:
    with sti.open(encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        if not r.fieldnames:
            return [], []
        rader: list[dict[str, str]] = []
        for row in r:
            rader.append({k: (row.get(k) or "").strip() for k in row.keys()})
        return [x.strip() if x else "" for x in r.fieldnames], rader


def reint_orgnr(verd: str) -> str | None:
    t = re.sub(r"\D", "", verd or "")
    if len(t) == 9:
        return t
    return None


def frå_excel_til_status(tek: str) -> str:
    """Gi grov kategori jamføringsvenleg med manuell_status."""
    if not (tek or "").strip():
        return "ukjent"
    s = tek.lower()
    if any(
        x in s
        for x in (
            "nedlagt",
            "oppløyst",
            "oppløys",
            "ikke aktiv",
            "ikkje aktiv",
            "opphørt",
            "opphort",
        )
    ):
        return "nedlagt"
    if "utmeld" in s or "utm." in s:
        return "utmeld"
    if any(
        x in s
        for x in (
            "medlem",
            "aktiv",
            "potensiell",
            "registrert",
        )
    ):
        return "medlem"
    if "ikkje" in s and ("aktuell" in s or "relevant" in s):
        return "ikkje_aktuell"
    return "ukjent"


def manuell_til_kategori(st: dict) -> str:
    if st.get("nedlagt") is True:
        return "nedlagt"
    nu = (st.get("nu") or "").strip().lower()
    if nu == "utmeld":
        return "utmeld"
    if nu in ("medlem", "potensiell_medlem"):
        return "medlem"
    if nu == "ikkje_aktuell":
        return "ikkje_aktuell"
    return "ukjent"


def gjett_orgnr_kolonne(h: list[str], rader: list[dict[str, str]]) -> str | None:
    best_k = None
    best_d = 0.0
    for k in h:
        vals = [reint_orgnr(row.get(k, "")) for row in rader if row]
        gode = sum(1 for v in vals if v)
        if not vals:
            continue
        brøk = gode / len(vals)
        lk = k.lower()
        if "org" in lk or "organisasjon" in lk:
            brøk += 0.15
        if brøk > best_d:
            best_d = brøk
            best_k = k
    if best_k and best_d >= 0.2:
        return best_k
    return None


def gjett_status_kolonne(h: list[str]) -> str | None:
    for k in h:
        lk = k.lower()
        if any(
            x in lk
            for x in (
                "status",
                "utmeld",
                "medlem",
                "nu",
                "tilstand",
                "situasjon",
                "aktiv",
                "nedlag",
            )
        ):
            return k
    return None


def last_register_orgnr(sti: Path) -> dict[str, dict[str, str]]:
    ut: dict[str, dict[str, str]] = {}
    if not sti.is_file():
        return ut
    with sti.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            o = reint_orgnr((row.get("orgnr") or "").strip())
            if not o:
                continue
            ut[o] = {k: (row.get(k) or "").strip() for k in row}
    return ut


def last_manuell(sti: Path) -> dict[str, dict]:
    if not sti.is_file():
        return {}
    rå = sti.read_text(encoding="utf-8").strip()
    if not rå or rå == "{}":
        return {}
    d = json.loads(rå)
    return d if isinstance(d, dict) else {}


def hovud() -> int:
    p = argparse.ArgumentParser(
        description="Jamfør NU Excel/CSV med docs/data/lag.csv og data/manuell_status.json.",
    )
    p.add_argument(
        "fil",
        type=Path,
        help="Sti til .xlsx eller .csv (CSV anbefalt utan openpyxl)",
    )
    p.add_argument(
        "--orgnr-kolonne",
        metavar="NAVN",
        help="Overskrift i arket for orgnr (elles: automatisk gjetting)",
    )
    p.add_argument(
        "--status-kolonne",
        metavar="NAVN",
        help="Overskrift for status (elles: automatisk gjetting)",
    )
    p.add_argument(
        "--ark",
        metavar="ARK",
        help="Vel Excel-ark: namn eller 0‑basert indeks",
    )
    p.add_argument(
        "--med-tom-manuell",
        action="store_true",
        help="List òg der Excel har status men manuell_status manglar oppføring for orgnr",
    )
    p.add_argument(
        "--max-utanfor-excel",
        type=int,
        default=40,
        metavar="N",
        help="Maks rader liste under «I lag.csv men ikkje i Excel» (0 = berre tal; default %(default)s)",
    )
    val = p.parse_args()

    ark: str | int | None = None
    if val.ark is not None:
        try:
            ark = int(val.ark)
        except ValueError:
            ark = val.ark

    sti = val.fil.resolve()
    org_kol: str | None = val.orgnr_kolonne
    status_kol: str | None = val.status_kolonne
    med_tom_manuell: bool = val.med_tom_manuell
    max_csv_utanom: int = max(0, val.max_utanfor_excel)

    if not sti.is_file():
        print(f"Finn ikkje {sti}", file=sys.stderr)
        return 1

    suff = sti.suffix.lower()
    if suff == ".csv":
        h, rader = les_csv(sti)
    elif suff in (".xlsx", ".xlsm"):
        if not krev_openpyxl():
            print(
                "Manglar openpyxl for .xlsx. Køyr: pip install openpyxl\n"
                "Eller lagre arket som CSV (UTF-8) og gje sti til .csv",
                file=sys.stderr,
            )
            return 1
        h, rader = les_xlsx(sti, ark)
    else:
        print(
            "Støtta format: .csv eller .xlsx — lagre frå Excel om nødvendig.",
            file=sys.stderr,
        )
        return 1

    if not rader:
        print("Ingen rader å lesa.", file=sys.stderr)
        return 1

    org_kol = org_kol or gjett_orgnr_kolonne(h, rader)
    if not org_kol:
        print(
            "Kunde ikkje finna orgnr‑kolonne. Set --orgnr-kolonne «Namn på overskrift».",
            file=sys.stderr,
        )
        return 1

    status_kol = status_kol or gjett_status_kolonne(h)

    pro = rot()
    reg_sti = pro / "docs" / "data" / "lag.csv"
    man_sti = pro / "data" / "manuell_status.json"
    if not reg_sti.is_file():
        print(
            f"Finn ikkje {reg_sti} — køyr `publiser_data_til_nettside.py` eller bruk `utdata/…` som kjelde.",
            file=sys.stderr,
        )
        return 1
    register = last_register_orgnr(reg_sti)
    manuell = last_manuell(man_sti)

    excel_orgn: set[str] = set()
    utan_gyldig_org: list[str] = []
    excel_ikkje_i_csv: list[tuple[str, str]] = []  # orgnr, ev. namn frå excel
    status_avvik: list[tuple[str, str, str, str, str]] = (
        []
    )  # orgnr, lagsnavn, excel_rå, excel_kat, manuell_kat

    lag_kol_candidates = [k for k in h if any(
        x in k.lower() for x in ("navn", "lag", "namn", "tit", "lokal")
    )]
    lag_kol = lag_kol_candidates[0] if lag_kol_candidates else None

    for row in rader:
        rad_txt = " | ".join(str(row.get(x, ""))[:40] for x in h[:4])
        o = reint_orgnr(row.get(org_kol, ""))
        if not o:
            if (row.get(org_kol, "") or "").strip():
                utan_gyldig_org.append(f"  (orgnr ugyldig) {rad_txt}")
            continue
        excel_orgn.add(o)
        xl_namn = ""
        if lag_kol:
            xl_namn = (row.get(lag_kol, "") or "").strip()
        if o not in register:
            excel_ikkje_i_csv.append((o, xl_namn))
            continue

        excel_rå = (row.get(status_kol, "") if status_kol else "").strip()
        ex_kat = frå_excel_til_status(excel_rå) if status_kol else "ukjent"
        st_m = manuell.get(o)
        if isinstance(st_m, dict):
            m_kat = manuell_til_kategori(st_m)
        else:
            m_kat = "ukjent"
        ln = (register.get(o) or {}).get("lagsnavn") or xl_namn or ""
        if not status_kol or ex_kat == "ukjent":
            continue
        if m_kat != "ukjent" and ex_kat != m_kat:
            status_avvik.append((o, ln, excel_rå, ex_kat, m_kat))
        elif med_tom_manuell and m_kat == "ukjent":
            status_avvik.append((o, ln, excel_rå, ex_kat, "(ikkje i manuell_status)"))

    i_csv_ikkje_excel: list[tuple[str, str]] = []
    for o, rad in sorted(register.items()):
        if o not in excel_orgn:
            i_csv_ikkje_excel.append((o, rad.get("lagsnavn", "")))

    print("— Jamføring: NU Excel / CSV ↔ lagsregisteret —\n")
    print(f"Fil: {sti}")
    print(f"Orgnr‑kolonne: «{org_kol}»", end="")
    if status_kol:
        print(f"    Status‑kolonne: «{status_kol}»")
    else:
        print("    Status‑kolonne: (ikkje funnen — ingen statusjamføring)")
    print(f"Excel/CSV‑rader med orgnr: {len(excel_orgn)}")
    print(f"lag.csv ({reg_sti.relative_to(pro)}): {len(register)} rader\n")

    print(f"** Rader utan gyldig 9‑siffer orgnr i «{org_kol}»:** {len(utan_gyldig_org)}")
    for x in utan_gyldig_org[:40]:
        print(x)
    if len(utan_gyldig_org) > 40:
        print(f"  … og {len(utan_gyldig_org) - 40} til")

    print(
        f"\n** Orgnr i Excel, men finst ikkje i lag.csv:** {len(excel_ikkje_i_csv)}"
    )
    for o9, xm in excel_ikkje_i_csv[:100]:
        print(f"  {o9}  {xm or '(utan namn i arket)'}")

    if status_kol:
        print(
            f"\n** Mogeleg status‑avvik (Excel vs data/manuell_status.json):** {len(status_avvik)}"
        )
        for o9, ln, xr, ek, mk in status_avvik[:200]:
            print(
                f"  {o9}  {ln[:60] if ln else ''}\n"
                f"      Excel «{xr}» → {ek}    register/manuell: {mk}"
            )
        if len(status_avvik) > 200:
            print(f"  … og {len(status_avvik) - 200} til")

    print(
        f"\n** I lag.csv, men ikkje funnen i Excel/CSV:** {len(i_csv_ikkje_excel)}"
    )
    print(
        "  (Stort tal er vanleg dersom Excel‑lista berre dekkjer NU‑lag, ikkje heile Brreg‑registeret.)\n"
    )
    if max_csv_utanom > 0:
        for o9, n in i_csv_ikkje_excel[:max_csv_utanom]:
            print(f"  {o9}  {n}")
        if len(i_csv_ikkje_excel) > max_csv_utanom:
            print(f"  … og {len(i_csv_ikkje_excel) - max_csv_utanom} til")
    else:
        print("  (bruk --max-utanfor-excel N for å lista rader, t.d. 80)")

    print(
        "\nMerk: Statusjamføring er grovt (tekstmønster). Sjå gjennom manuelt før du endrar json."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(hovud())