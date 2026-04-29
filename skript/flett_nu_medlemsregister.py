#!/usr/bin/env python3
"""
Flett NU medlemsregister (Excel) inn i lag.csv → nu_mr_status, nu_mr_overordna, nu_mr_orgtype.

Rader der Excel har **organisasjonsnummer** som **ikkje finst** etter innhent frå Brreg
(felt i `lag.csv` som kjem frå søkjefragment‑treff) vert også **lagde til**, med
kjelde «nu_mr_orgnr_ikkje_i_brreg» at alle NU‑«Aktive» kan teljast på nettsida.
Vel ``--ikkje-vedlegg-orgnr-utanfor-brreg`` om du ikkje vil ha dei.

Rader i Excel **utan organisasjonsnummer** som ikkje vert koplast til ei Brreg‑rad, vert som standard
**lagt til** i CSV med tom orgnr og kjelde «nu_mr_utan_orgnr», slik at dei visast på nettsida.
Vel ``--ikkje-vedlegg-utan-orgnr`` om du ikkje vil ha slike tilleggsrader.

Sjå dokumentasjon i modul‑doc‑strengen øvst på fila (hjelp).

Døme::

  pip install openpyxl
  python3 skript/flett_nu_medlemsregister.py /sti/NU/oversikt.xlsx

  python3 skript/flett_nu_medlemsregister.py oversikt.xlsx \\
      --csv docs/data/lag.csv --synk-manuell
"""
from __future__ import annotations

import argparse
import csv
import difflib
import json
import re
import shutil
import sys
from datetime import date
from pathlib import Path


def rot() -> Path:
    return Path(__file__).resolve().parent.parent


def norm_stad(s: str) -> str:
    return " ".join((s or "").strip().split()).lower()


def reint_orgnr(verd: str) -> str | None:
    t = re.sub(r"\D", "", verd or "")
    if len(t) == 9:
        return t
    return None


def norm_namn(namn: str) -> str:
    t = namn.strip().lower()
    t = t.replace("fril.", "frilynde")
    t = re.sub(r"\bu\s*/\s*l\b", "ul", t)
    t = re.sub(r"[^a-zæøå0-9]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def finn_kolonne_for(
    headers: list[str],
    *må_alle_liggje: str,
) -> str | None:
    """Fyrste overskrift der alle små bokstavar‑delstrengar finst i overskrifta."""
    best: str | None = None
    best_len = 999
    for h in headers:
        sl = h.strip().lower()
        if all(m in sl for m in må_alle_liggje):
            if len(h) < best_len:
                best_len = len(h)
                best = h
    return best


def les_excel(sti: Path, ark: str | None) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(sti, read_only=True, data_only=True)
    ws = wb[ark] if ark and ark in wb.sheetnames else wb.active
    rår = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rår:
        return [], []
    råh: list[str] = []
    tell: dict[str, int] = {}
    for c in rår[0]:
        s = str(c).strip() if c is not None else ""
        grunn = s or f"kol_{len(råh)}"
        n = tell.get(grunn, 0)
        tell[grunn] = n + 1
        key = grunn if n == 0 else f"{grunn}_{n}"
        råh.append(key)
    ut: list[dict[str, str]] = []
    for rad in rår[1:]:
        d: dict[str, str] = {}
        tom = True
        for i, k in enumerate(råh):
            v = rad[i] if i < len(rad) else None
            if v is None:
                s = ""
            elif isinstance(v, float) and v == int(v):
                s = str(int(v))
            else:
                s = str(v).strip()
            if s:
                tom = False
            d[k] = s
        if not tom:
            ut.append(d)
    return råh, ut


def les_csv_rader(sti: Path) -> tuple[list[str], list[dict[str, str]]]:
    with sti.open(encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        if not r.fieldnames:
            return [], []
        fn = list(r.fieldnames)
        rader = []
        for row in r:
            rader.append({k: (row.get(k) or "").strip() for k in fn})
        return fn, rader


def skriv_csv(sti: Path, felt: list[str], rader: list[dict[str, str]]) -> None:
    sti.parent.mkdir(parents=True, exist_ok=True)
    with sti.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=felt, extrasaction="ignore")
        w.writeheader()
        for rad in rader:
            w.writerow({k: rad.get(k, "") for k in felt})


def lik_namn(a: str, b: str) -> float:
    na, nb = norm_namn(a), norm_namn(b)
    if not na or not nb:
        return 0.0
    return difflib.SequenceMatcher(None, na, nb).ratio()


KJELDE_NU_UTAN_ORGNR = "nu_mr_utan_orgnr"
KJELDE_NU_ORGNR_IKKJE_BRREG = "nu_mr_orgnr_ikkje_i_brreg"


def liste_frå_nu_orgtype(orgtype_tekst: str | None) -> str:
    """Ungdom* vs Bygd* i «liste», etter NU sin organisasjonstype."""
    kt = (orgtype_tekst or "").strip().lower()
    if kt and "fylke" in kt:
        return "bygdelag"
    return "ungdomslag"


def namn_komm_fylk_lik(
    navn: str, komm: str, fylk: str, rad: dict[str, str]
) -> bool:
    return (
        norm_namn(navn) == norm_namn(rad.get("lagsnavn", ""))
        and norm_stad(komm) == norm_stad(rad.get("kommune", ""))
        and norm_stad(fylk) == norm_stad(rad.get("fylke", ""))
    )


def finst_lik_rade_i_register(
    navn: str, komm: str, fylk: str, rader: list[dict[str, str]]
) -> bool:
    """True dersom same namn+kommune+fylke alt finst (orgnr eller tom)."""
    for rad in rader:
        if namn_komm_fylk_lik(navn, komm, fylk, rad):
            return True
    return False


def hovud() -> int:
    p = argparse.ArgumentParser(
        description="Flett NU medlemsregister (Excel) inn i lag.csv",
    )
    p.add_argument("excel", type=Path, help="Sti til .xlsx")
    p.add_argument(
        "--csv",
        type=Path,
        default=None,
        help="lag.csv (default: docs/data/lag.csv under prosjektrot)",
    )
    p.add_argument("--ark", default=None, help="Excel-arknamn (default: aktivt ark)")
    p.add_argument(
        "--synk-manuell",
        action="store_true",
        help="Oppdater data/manuell_status.json frå NU-status (berre matchende orgnr)",
    )
    p.add_argument(
        "--terskel-namn",
        type=float,
        default=0.86,
        help="Min. likskap namn når orgnr manglar (default %(default)s)",
    )
    p.add_argument(
        "--ikkje-vedlegg-utan-orgnr",
        action="store_true",
        help=(
            "Ikkje legg til nye CSV‑rader for NU‑Excel utan orgnr som ikkje allereie "
            "finst i registeret (vanlegvis vil du ha dei med)."
        ),
    )
    p.add_argument(
        "--ikkje-vedlegg-orgnr-utanfor-brreg",
        action="store_true",
        help=(
            "Ikkje legg til NU‑Excel‑rader med orgnr som ikkje kom med i Brreg‑innhenting "
            "(standard: legg dei til — trengst for at alle Aktive tel med)."
        ),
    )
    val = p.parse_args()

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("Køyr: pip install openpyxl", file=sys.stderr)
        return 1

    pro = rot()
    csv_sti = val.csv or (pro / "docs" / "data" / "lag.csv")
    csv_sti = csv_sti.resolve()
    if not csv_sti.is_file():
        print(f"Finn ikkje {csv_sti}", file=sys.stderr)
        return 1
    ex_sti = val.excel.resolve()
    if not ex_sti.is_file():
        print(f"Finn ikkje {ex_sti}", file=sys.stderr)
        return 1

    h, ex_rader = les_excel(ex_sti, val.ark)
    if not ex_rader:
        print("Tomt Excel-ark.", file=sys.stderr)
        return 1

    k_namn = finn_kolonne_for(h, "namn") or finn_kolonne_for(h, "lag")
    k_stat = finn_kolonne_for(h, "status")
    k_komm = finn_kolonne_for(h, "kommun")
    k_fylk = finn_kolonne_for(h, "fylke")
    k_org = finn_kolonne_for(h, "organisasjon", "nummer")
    k_over = finn_kolonne_for(h, "ovanstå", "organ") or finn_kolonne_for(
        h, "overord"
    )
    k_type = finn_kolonne_for(h, "organisasjonstype") or finn_kolonne_for(h, "type")

    mang = [x for x, y in [
        ("namn", k_namn),
        ("status", k_stat),
        ("kommune", k_komm),
        ("fylke", k_fylk),
    ] if not y]
    if mang:
        print(
            f"Kunde ikkje finna kolonne for: {', '.join(mang)}. Overskrifter: {h}",
            file=sys.stderr,
        )
        return 1

    assert k_namn and k_stat and k_komm and k_fylk

    excel_per_org: dict[str, dict[str, str]] = {}
    ex_utan_org: list[dict[str, str]] = []
    for ex in ex_rader:
        o = reint_orgnr(ex.get(k_org or "", ""))
        if o:
            excel_per_org[o] = ex
        else:
            ex_utan_org.append(ex)

    felt, rader = les_csv_rader(csv_sti)
    nye_k = ["nu_mr_status", "nu_mr_overordna", "nu_mr_orgtype"]
    mang_k = [k for k in nye_k if k not in felt]
    if mang_k:
        if "liste" in felt:
            ins = felt.index("liste") + 1
            for j, k in enumerate(mang_k):
                felt.insert(ins + j, k)
        else:
            felt.extend(mang_k)

    manuell_oppdatert: dict[str, dict] = {}
    man_sti = pro / "data" / "manuell_status.json"
    if val.synk_manuell and man_sti.is_file() and man_sti.read_text().strip():
        manuell_oppdatert = json.loads(man_sti.read_text(encoding="utf-8"))
        if not isinstance(manuell_oppdatert, dict):
            manuell_oppdatert = {}

    brukt_utan_org: set[int] = set()
    treff_org = 0
    treff_namn = 0
    ikkje_nu = 0

    def fyll_nu_felt(rad: dict[str, str], ex: dict[str, str]) -> None:
        rad["nu_mr_status"] = (ex.get(k_stat, "") or "").strip()
        rad["nu_mr_overordna"] = (
            (ex.get(k_over, "") or "").strip() if k_over else ""
        )
        rad["nu_mr_orgtype"] = (ex.get(k_type, "") or "").strip() if k_type else ""

    def prøv_namn_kf(rad: dict[str, str]) -> bool:
        """Treff i ex_utan_org på namn + kommune + fylke. Retnar True ved treff."""
        lnav = rad.get("lagsnavn", "")
        rk = norm_stad(rad.get("kommune", ""))
        rf = norm_stad(rad.get("fylke", ""))
        best_i = -1
        best_sc = 0.0
        for ei, ex in enumerate(ex_utan_org):
            if ei in brukt_utan_org:
                continue
            ek = norm_stad(ex.get(k_komm, ""))
            ef = norm_stad(ex.get(k_fylk, ""))
            if rk and ek and rk != ek:
                continue
            if rf and ef and rf != ef:
                continue
            sc = lik_namn(lnav, ex.get(k_namn, ""))
            if sc > best_sc:
                best_sc = sc
                best_i = ei
        if best_i >= 0 and best_sc >= val.terskel_namn:
            fyll_nu_felt(rad, ex_utan_org[best_i])
            brukt_utan_org.add(best_i)
            return True
        return False

    def excel_til_manuell(sttxt: str) -> dict | None:
        s = sttxt.strip().lower()
        if not s:
            return None
        if s == "aktiv":
            return {"nu": "medlem"}
        if s == "inaktiv":
            return {"nu": "inaktiv_medlem"}
        if s == "nedlagt":
            return {"nedlagt": True}
        if "utmeld" in s or s == "utmeld":
            return {"nu": "utmeld"}
        return None

    for rad in rader:
        org = reint_orgnr(rad.get("orgnr", ""))
        if org and org in excel_per_org:
            fyll_nu_felt(rad, excel_per_org[org])
            treff_org += 1
            if val.synk_manuell:
                ku = excel_til_manuell(excel_per_org[org].get(k_stat, "") or "")
                if ku is not None:
                    nå = dict(manuell_oppdatert.get(org, {}) or {})
                    if isinstance(nå, dict):
                        if "nu" in ku:
                            nå["nu"] = ku["nu"]
                        if ku.get("nedlagt"):
                            nå["nedlagt"] = True
                        manuell_oppdatert[org] = nå
            continue

        # Ingen orgnr‑treff i Excel eller ingen orgnr i CSV: namn + kommune + fylke
        if prøv_namn_kf(rad):
            treff_namn += 1
            if val.synk_manuell and org:
                ku = excel_til_manuell(rad.get("nu_mr_status", "") or "")
                if ku is not None:
                    nå = dict(manuell_oppdatert.get(org, {}) or {})
                    if isinstance(nå, dict):
                        if "nu" in ku:
                            nå["nu"] = ku["nu"]
                        if ku.get("nedlagt"):
                            nå["nedlagt"] = True
                        manuell_oppdatert[org] = nå
            continue

        for k in nye_k:
            rad.setdefault(k, "")
        ikkje_nu += 1

    # Fyll tom for rader utan match
    for rad in rader:
        for k in nye_k:
            rad.setdefault(k, "")

    dags_i_dag = date.today().isoformat()

    ubrukt_indexes = [i for i in range(len(ex_utan_org)) if i not in brukt_utan_org]
    vedlegg_utan_org = 0
    hop_duplikat = 0
    hop_tomme_namn = 0
    if not val.ikkje_vedlegg_utan_orgnr:
        for ei in ubrukt_indexes:
            ex = ex_utan_org[ei]
            navn = (ex.get(k_namn) or "").strip()
            komm = (ex.get(k_komm) or "").strip()
            fylk = (ex.get(k_fylk) or "").strip()
            if not navn:
                hop_tomme_namn += 1
                continue
            if finst_lik_rade_i_register(navn, komm, fylk, rader):
                hop_duplikat += 1
                continue
            ny: dict[str, str] = {k: "" for k in felt}
            ny["lagsnavn"] = navn
            ny["orgnr"] = ""
            ny["kommune"] = komm
            ny["fylke"] = fylk
            ny["kjelde_type"] = KJELDE_NU_UTAN_ORGNR
            ny["kjelde_url"] = ""
            ny["liste"] = liste_frå_nu_orgtype(ex.get(k_type) if k_type else None)
            ny["henta_dato"] = dags_i_dag
            fyll_nu_felt(ny, ex)
            rader.append(ny)
            vedlegg_utan_org += 1

    orgnr_i_register = {o for o in (reint_orgnr(r.get("orgnr", "")) for r in rader) if o}
    vedlegg_orgnr_utanfor = 0
    if not val.ikkje_vedlegg_orgnr_utanfor_brreg:
        for org9 in sorted(excel_per_org.keys()):
            if org9 in orgnr_i_register:
                continue
            ex = excel_per_org[org9]
            navn = (ex.get(k_namn) or "").strip()
            if not navn:
                continue
            ny = {k: "" for k in felt}
            ny["lagsnavn"] = navn
            ny["orgnr"] = org9
            ny["kommune"] = (ex.get(k_komm) or "").strip()
            ny["fylke"] = (ex.get(k_fylk) or "").strip()
            ny["kjelde_type"] = KJELDE_NU_ORGNR_IKKJE_BRREG
            ny["kjelde_url"] = ""
            ny["liste"] = liste_frå_nu_orgtype(ex.get(k_type) if k_type else None)
            ny["henta_dato"] = dags_i_dag
            fyll_nu_felt(ny, ex)
            rader.append(ny)
            orgnr_i_register.add(org9)
            vedlegg_orgnr_utanfor += 1

    # Backup
    backup = csv_sti.with_suffix(".csv.før_nu_medlemsregister.bak")
    try:
        shutil.copy2(csv_sti, backup)
    except OSError:
        pass

    skriv_csv(csv_sti, felt, rader)
    print(
        f"Skreiv {csv_sti.relative_to(pro)} — treff orgnr: {treff_org}, "
        f"treff namn+k+f (Excel utan orgnr eller CSV utan orgnr): {treff_namn}, "
        f"registerrader utan NU‑treff: {ikkje_nu}",
        file=sys.stderr,
    )
    if not val.ikkje_vedlegg_utan_orgnr:
        print(
            f"  Nye rader lagt til (kun NU, utan orgnr i eksisterande register): "
            f"{vedlegg_utan_org}",
            file=sys.stderr,
        )
        if hop_duplikat or hop_tomme_namn:
            print(
                f"  (hoppa over: tomme namn {hop_tomme_namn}, "
                f"samme lag finst alt i CSV {hop_duplikat})",
                file=sys.stderr,
            )

    if not val.ikkje_vedlegg_orgnr_utanfor_brreg:
        print(
            f"  Nye rader lagt til (orgnr frå NU, ikkje i Brreg‑innhenting før): "
            f"{vedlegg_orgnr_utanfor}",
            file=sys.stderr,
        )

    excel_ikkje_brreg = [
        o for o in excel_per_org if o not in {reint_orgnr(r.get("orgnr", "")) for r in rader if reint_orgnr(r.get("orgnr", ""))}
    ]
    if excel_ikkje_brreg[:20]:
        print(
            "\n** Døme på orgnr i NU‑Excel som ikkje finst i lag.csv (fyrste 20):**",
            file=sys.stderr,
        )
        for o in excel_ikkje_brreg[:20]:
            print(f"  {o}", file=sys.stderr)

    if val.ikkje_vedlegg_utan_orgnr:
        ubr_utv = [i for i in range(len(ex_utan_org)) if i not in brukt_utan_org]
        if ubr_utv:
            print(
                f"\n** {len(ubr_utv)} NU‑rader utan orgnr vert ikkje lagde til** "
                "(du brukte --ikkje-vedlegg-utan-orgnr). "
                "Køyr på nytt utan flagget om du vil inn med alt som manglar.",
                file=sys.stderr,
            )

    if val.synk_manuell:
        man_sti.parent.mkdir(parents=True, exist_ok=True)
        man_sti.write_text(
            json.dumps(manuell_oppdatert, ensure_ascii=False, indent=2, sort_keys=True)
            + "\n",
            encoding="utf-8",
        )
        print(
            f"Oppdaterte {man_sti.relative_to(pro)} (synk frå NU-status).",
            file=sys.stderr,
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(hovud())
